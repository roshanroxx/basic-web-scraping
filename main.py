
from ast import Global
from re import X, search
from unicodedata import name
from xml.sax.handler import property_declaration_handler
import openpyxl
from email.utils import decode_rfc2231
from selenium import webdriver as wb
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook
import csv
import pandas as pd
import time


webD = wb.Chrome(executable_path="C:\\Users\\dell\\Downloads\\chromedriver_win32\\chromedriver.exe")
webD.get("https://www.classcentral.com/")
webD.maximize_window()

#//*[@id="page-search"]/div[1]/header/div[1]/nav/div[4]/form/fieldset/input[1]
count = 0
def find(word,count) :
    
    
    search = webD.find_element(By.NAME,'q')
    
      
    search.send_keys(Keys.BACKSPACE*count)



    search.send_keys(word + Keys.RETURN)

    


    free = webD.find_element(By.XPATH,'//*[@id="page-search"]/div[1]/div[2]/div[4]/div/ol/li[1]/ol/li[2]/fieldset/label/span[1]')
    free.click()
    rdoc1 = webD.find_elements(By.XPATH,'//*[@id="page-search"]/div[1]/div[2]/div[5]/ol/li[*]/div[1]/div[1]/div[2]/a[1]/h2')
    rdoc2 = webD.find_elements(By.XPATH,'//*[@id="page-search"]/div[1]/div[2]/div[5]/ol/li[*]/div[1]/div[1]/div[2]/a[2]/span[2]')
    rdoc3 = webD.find_elements(By.XPATH,'//*[@id="page-search"]/div[1]/div[2]/div[5]/ol/li[*]/div[1]/div[1]/div[2]/a[1]')
    doc1 = []
    doc2 = []
    doc3 = []
#//*[@id="page-search"]/div[1]/div[2]/div[5]/ol/li[3]/div[1]/div[1]/div[2]/a[1]
#/html/body/div[1]/div[2]/div[5]/ol/li[2]/div[1]/div[1]/div[2]/a[1]
    for i in rdoc1 :
        doc1.append(i.text)
    for i in rdoc2 :
        doc2.append(i.text)
    
    for i in rdoc3 :
        doc3.append(i.get_attribute('href'))
    
    doc = zip(doc1 , doc3)


    count


    return doc



def csv(doc,name) :
    df = pd.DataFrame(doc)
    df.to_csv('{}.csv'.format(name), index=False)




list1 = ['ai','machine learning','frontend','fullstack','backend']


for i in range(len(list1)) :
    if list1[i]!=list1[0] :
         count = len(list1[i-1])

    doc = list(find(list1[i],count))

    csv(doc,list1[i])
    time.sleep(1)
    





""""
wb = Workbook()
sh1 = wb.active

for i in list(doc) :
    sh1.append(i)

wb.save('final.xlsx')


with open('test.csv', 'wb', newline='') as myfile:
    csv_writer = csv.writer(myfile, delimiter='|', quotechar='"', quoting=csv.QUOTE_MINIMAL)
 #   csv_writer = csv.writer(myfile)
    for row in sh1.iter_rows() :
        csv_writer.writerow(cell.value for cell in row)

"""
#for row in sh.iter_rows(): # generator; was sh.rows
#       csv_writer.writerow([cell.value for cell in row])
#//*[@id="page-search"]/div[1]/div[2]/div[5]/ol/li[2]/div[1]/div[1]/div[2]/a[2]/span[2]

#//*[@id="page-search"]/div[1]/div[2]/div[5]/ol/li[3]/div[1]/div[1]/div[2]/a[2]/span[2]

"""
machine learning
ai
frontend
backend
fullstack
"""
